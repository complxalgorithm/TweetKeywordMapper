"""

@Name: data.py
@Author: Stephen Sanders <https://stephensanders.me>
@Description: Contains most functions that are imperative to running the program. This includes the 
              file_interact() function, directory and file related functions, and other smaller functions.
@Requirements: pandas

"""


# import libraries
import pandas as pd
import os
import time
import csv

# change pandas options so that all results will be displayed
# also tells pandas to show entire Tweet ID integers without truncation
pd.set_option('display.max_rows', 999)
pd.set_option('display.max_columns', 999)
pd.set_option('display.width', 999)
pd.set_option('display.max_colwidth', 999)
pd.set_option('display.float_format', lambda x: f'%.{0 if x.is_integer() else 1}f' % x)


"""
# define get_user_file() function - get a file from the user from which to read data,
# then return valid file and its file extension
"""
def get_user_file(default_file):
    # ask user to enter their file, or use the default file
    user_file = input('Enter CSV or XlSX file from your current directory, or hit enter to use default file: ') or default_file
    
    # automatically return file and its extension if the input is the same as the default file
    if user_file == default_file:
        file_ext = default_file.split('.')[-1]
        return default_file, file_ext
    
    # get extension of input file
    file_ext = user_file.split('.')[-1]
    
    # validate that file exists and that the file is a CSV or XLSX file
    while os.path.exists(user_file) == False or (file_ext != 'csv' and file_ext != 'xlsx'):
        # display error if input doesn't exist
        if os.path.exists(user_file) == False:
            print(f'{user_file} does not exist in your current directory.')
        
        # display error if input is not a CSV or Excel file
        else:
            # display error if input isn't a CSV file
            if file_ext != 'csv':
                print(f'{user_file} is not a CSV file.')

            # display error if input isn't an Excel file
            else:
                print(f'{user_file} is not an XLSX file.')
            
        # ask user for file again, and use default file if input is left blank
        user_file = input('Enter file from your current directory, or hit enter to use default file: ') or default
        
        # get extension of new input file
        file_ext = user_file.split('.')[-1]
    
    # return the file and file extension to the parent function
    return user_file, file_ext


"""
# define get_file_contents() function - reads contents of file depending on the file extension
# returns contents df and list of fields in file
"""
def get_file_contents_fields(file, file_ext):
    # get current working directory (i.e., the project directory)
    cwd = os.getcwd()

    # set file path
    file_path = os.path.join(cwd, file)
    
    # get contents from file
    if file_ext == 'csv':
        # extract contents of CSV file
        contents = pd.read_csv(file_path, header=0, sep=',')
    
    # get contents from XLSX file
    else:
        # get current working directory (i.e., the project directory)
        cwd = os.getcwd()

        # set file path
        file_path = os.path.join(cwd, file)
        
        # extract contents of XLSX file
        contents = pd.read_excel(file_path, header=0, engine='openpyxl')
        
    # get fields from file
    fields = [f for f in contents]
    
    # return file contents and field names
    return contents, fields


"""
# define get_shp_directory() function - asks user to specify which directory within workspace contains shapefile
"""
def get_shp_directory(ws):
    # create empty dirs_list to hold all directories in project
    dirs_list = []
    
    # go through every file/dir that is within the workspace
    for file in os.listdir(ws):
        # set path to current file/dir
        path = os.path.join(ws, file)
        
        # try to split the path using a period in order to identify which ones have a file extension
        try_dir = path.split('.')
        
        # run this if file path was not split
        if len(try_dir) == 1:
            # make sure file is a directory
            if os.path.isdir(path) == True:
                # make sure directory doesn't end with an underscore
                if path[-1] != '_':
                    # add this path to dirs_list
                    dirs_list.append(path)
    
    # append workspace to dirs_list
    dirs_list.append(ws)
    
    # generate key values for each dir in dir_list to add to dirs dict
    nums = [str(i) for i in range(1, len(dirs_list)+1)]
    
    # create dictionary of keyed directory values
    dirs = dict(zip(nums, dirs_list))
    
    # display menu of possible directories
    print()
    for d in dirs:
        print(f'{d}: {dirs[d]}')
    print()
    
    # ask user to choose which directory their US states shapefile is in
    user_dir = input('Choose which directory your shapefile is in using its number: ')
    
    # validate that user_dir exists
    while user_dir not in dirs:
        # display an error, then pause program for half a second
        print('That is not a directory in your project.')
        time.sleep(0.5)
        
        # ask user again to choose which directory their US states shapefile is in
        user_dir = input('Choose which directory your shapefile is in using its number: ')
    
    # return directory to parent function
    return dirs[user_dir]


"""
# define get_shapefile() function - returns shapefile to be used by GeoPandas to map each state's Tweet counts
"""
def get_shapefile(ws):
    # get directory that contains US states shapefile
    user_dir = get_shp_directory(ws)
    
    # create empty shp_list to hold all shapefiles in user_dir
    shp_list = []
    
    # go through every file/dir that is within user_dir
    for file in os.listdir(user_dir):
        # set path to current file/dir
        path = os.path.join(user_dir, file)
        
        # append current file to shp_list if file has a .shp extension
        if file.split('.')[-1] == 'shp':
            shp_list.append(path)
    
    # validate that the directory contains shapefiles
    while len(shp_list) == 0:
        # tell user the directory they picked doesn't have any shapefiles,
        # then pause program for half a second
        print(f'{user_dir} has no shapefiles. Please choose a different directory.')
        time.sleep(0.5)
        
        # ask user for another directory
        user_dir = get_shp_directory(ws)
        
        # create empty shp_list to hold all shapefiles in user_dir
        shp_list = []
    
        # go through every file/dir that is within user_dir
        for file in os.listdir(user_dir):
            # set path to current file/dir
            path = os.path.join(user_dir, file)
        
            # append current file to shp_list if file has a .shp extension
            if file.split('.')[-1] == 'shp':
                shp_list.append(path)
    
    # generate key values for each shp in shp_list to add to dirs dict
    nums = [str(i) for i in range(1, len(shp_list)+1)]
    
    # create dictionary of keyed shapefile values
    shps = dict(zip(nums, shp_list))
    
    # display menu of shapefiles
    print()
    for s in shps:
        print(f'{s}: {shps[s]}')
    print()
    
    # ask user to choose which directory their US states shapefile is in
    user_shp = input('Choose your US states shapefile using its number: ')
    
    # validate that user_dir exists
    while user_shp not in shps:
        # display an error
        print(f'{user_shp} is not a shapefile.')
        
        # ask user again to choose which directory their US states shapefile is in
        user_shp = input('Choose your US states shapefile using its number: ')
    
    # return the shapefile and directory to the parent function
    return shps[user_shp], user_dir


"""
# define get_values() function - get unique values of selected field, and return list of unique values
# and the selected field
"""
def get_unique_values(contents, fields):
    # display menu of fields that were found in file
    print()
    for i, field in enumerate(fields, start=1):
        print(f'{i}. {field}')
    print()

    # initialize empty lists of unique keywords from file and counts of each unique keyword in keywords list
    values = []
    counts = []

    # ask user to identify the keywords field from the menu
    field = input('Enter the field you would like to tally from the menu: ')

    # validate that field is in fields list
    while field not in fields:
        # tell user that field doesn't exist, then pause program for a second
        print(f'{field} field does not exist.')
        time.sleep(1)

        # ask user again for keyword field
        field = input('Enter the field you would like to tally from the menu: ')

    print()

    # initialize list to store keyword values object containing values from keyword_field
    data = []

    # add data values/object of keyword_field to keyword_data list
    data.append(contents[field])

    # iterate through each keyword result object and add unique values to keyword list
    for ob in data:
        for word in ob:
            # add keyword to keywords list if it isn't already in it
            if word not in values:
                values.append(word)
    
    # return list of values and field lists to parent function
    return values, field


"""
# define get_field_indexes_tweet_ids() function - determines the location of each relevant field using the user's input
# also generates a list of Tweet IDs that are in the file to make sure duplicate Tweets aren't being added to the file
# returns the list of Tweet IDs and the indexes of the state, tweet_id, and keyword fields
"""
def get_field_indexes_tweet_ids(contents, fields):
    # display menu of fields that were found in file
    print()
    for i, field in enumerate(fields, start=1):
        print(f'{i}. {field}')
    print()

    # ask user to enter field that has states
    state_field = input('Enter the field that contains names of states: ')

    # validate that specified state field is in fields list
    while state_field not in fields:
        # tell user that state field isn't in the fields list
        print(f'{state_field} does not exist.')

        # tell user to enter the states field again
        state_field = input('Enter the field that contains the names of states: ')

    # ask user to enter field that has Tweet IDs
    id_field = input('Enter the field that contains Tweet IDs: ')

    # validate that specified ids field is in fields list and is not the same as states field
    while id_field not in fields or id_field == state_field:
        # tell user if field does not exist
        if id_field not in fields:
            print(f'{id_field} does not exist.')

        # tell user if field is the same as state field
        if id_field == state_field:
            print(f'That is the same as {state_field}.')

        # tell user to enter the ids field again
        id_field = input('Enter the field that contains Tweet IDs: ')

    # ask user to enter field that contains keywords
    keyword_field = input('Enter the field that contains keywords: ')

    # validate that specified keywords field is in fields list and is not the same as states or ids fields
    while keyword_field not in fields or keyword_field == state_field or keyword_field == id_field:
        # tell user if field does not exist
        if keyword_field not in fields:
            print(f'{keyword_field} does not exist.')

        # tell user if field is the same as state field
        if keyword_field == state_field:
            print(f'That is the same as {state_field}.')

        # tell user if field is the same as id field
        if keyword_field == id_field:
            print(f'That is the same as {id_field}.')

        # tell user to enter the ids field again
        keyword_field = input('Enter the field that contains keywords: ')

    # get list of tweet ids from file using user id_field
    tweets = contents[id_field].tolist()

    # determine indexes of each of the specified fields in the fields list
    state_index = fields.index(state_field)
    id_index = fields.index(id_field)
    keyword_index = fields.index(keyword_field)
    
    # return Tweet IDs list and indexes to parent function
    return tweets, state_index, id_index, keyword_index


"""
# define file_interact() function - appends to, writes to, or reads contents of user's CSV/XLSX file
# optional parameters:
#   - mode         -> default to append         : values -> 'a', 'w', 'r'
#   - checkKeyword -> default to False          : values -> True, False
#   - function     -> default to empty string   : values -> 'counts', 'search', 'read'/blank value
"""
def file_interact(data, file, file_ext, workspace, mode='a', checkKeyword=False, function=''):
    # write to or append data to csv file
    if mode == 'a' or mode == 'w':
        # run when file is a CSV
        if file_ext == 'csv':
            # this code makes sure that a file in append mode will append the first line to a newline
            # it will only run if the csv file already exists
            # CREDIT: tdelaney from stack overflow
            # -> <https://stackoverflow.com/questions/64921222/csv-writer-adds-the-first-line-to-the-last-cell>
            if mode == 'a' and os.path.exists(file):
                with open(file, 'a+b') as f:
                    f.seek(-1, 2)

                    if f.read(1) != b"\n":
                        f.write(b"\r\n")
            
            # write data to the file, separating each unique value with a comma
            with open(file, mode=mode, newline='') as csv_write:
                csvwriter = csv.writer(csv_write, delimiter=',')
                csvwriter.writerow(data)

        # run when file is an Excel file
        else:
            # 
            if os.path.exists(file):
                # import load_workbook function from openpyxl
                from openpyxl import load_workbook
                
                # load the file
                wb = load_workbook(file)
                
                # get current working sheet
                page = wb.active
                
                # append data to sheet
                page.append(data)
                
                # save the updated file
                wb.save(filename=file)
            
            # create XLSX file when it currently doesn't exist
            else:
                # create df using data
                # in this case data will be the three default field names: Tweet_ID, Keyword, and State
                df = pd.DataFrame([data])
                
                # remove the numbered fields that are generated at the top
                # so that above default field names are the true column names of the file
                df.columns = df.iloc[0, :]
                df.drop(df.index[0], inplace=True)
                
                # write default field names to file
                with pd.ExcelWriter(file) as writer:
                    df.to_excel(writer, index=False)

        # return to the parent function
        return
    
    # read contents of file
    else:
        # extract contents df and fields list from file
        contents, fields = get_file_contents_fields(file, file_ext)
        
        # run this when functionality is being called from counts module
        if function == 'counts':
            # get list of unique keywords and keyword field from file,
            # then return both values to parent function
            values, field = get_unique_values(contents, fields)
            
            # return values, field, file contents df to parent function
            return values, field, contents
        
        # run this functionality is being called from search module
        elif function == 'search':
            # get ids and places lists from data list
            ids = data[0]
            places = data[1]
            
            # create tweet_data dict using ids and places lists
            tweet_data = dict(zip(ids, places))

            # get locations of data Tweet IDs, states, and keywords in the file using their indexes
            # in the user's file
            # also get a list of Tweet IDs that are already in the user's file
            # return this data plus the tweet_data once complete
            tweets, state_index, id_index, keyword_index = get_field_indexes_tweet_ids(contents, fields)
            
            return tweets, tweet_data, state_index, id_index, keyword_index
        
        # run this when functionality is being called from read module
        else:
            # initialize states and ids lists to store states and Tweet IDs from file
            states = []
            ids = []

            # display menu of fields that were found in file
            print()
            for i, field in enumerate(fields, start=1):
                print(f'{i}. {field}')
            print()

            # ask user to enter the state and tweet id fields from the menu
            state_field = input('Enter the field that contains names of states: ')

            # validate that the state field is in the fields list
            while state_field not in fields:
                print(f'{state_field} does not exist.')

                # ask user to enter the state field again
                state_field = input('Enter the field that contains names of states: ')

            # ask user to enter the tweet id field from the menu
            id_field = input('Enter the field that contains Tweet IDs: ')

            # validate that the id field is in the fields list
            while id_field not in fields or id_field == state_field:
                # tell user that the field is not in the fields list
                if id_field not in fields:
                    print(f'{id_field} does not exist.')

                # tell user that their id field is the same as their states field
                elif id_field == state_field:
                    print(f'{id_field} is the same as your states field.')

                # tell user to enter the id field again
                id_field = input('Enter the field that contains Tweet IDs: ')

            # intialize state_data and id_data lists to store values of state_field and id_field
            state_data = []
            id_data = []

            # run keyword filtering related code if function parameter indicates to do so
            if checkKeyword == True:
                # ask user if they'd like to only map tweets that have a certain keyword
                if_keyword = input('Would you like to import Tweets that have a certain keyword? (Y or N) ')

                # validate that input is Yes/Y or No/N
                while if_keyword.title() != 'Yes' and if_keyword.upper() != 'Y' and if_keyword.title() != 'No' and if_keyword.upper() != 'N':
                    # display error if input isn't Yes/Y or No/N
                    print(f'{if_keyword} is not a valid answer. Please try again.')

                    # ask user again
                    if_keyword = input('Would you like to import Tweets that have a certain keyword? (Y or N) ')

                # run this code if the user wants to only include Tweets with a certain keyword from the file
                if if_keyword.title() == 'Yes' or if_keyword.upper() == 'Y':
                    # initialize empty list of unique keywords from file
                    keywords = []

                    # ask user to identify the keywords field from the menu
                    keyword_field = input('Enter the field that contains keywords: ')

                    # validate that field is in fields list
                    while keyword_field not in fields or keyword_field == state_field or keyword_field == id_field:
                        # display that the keyword field is the same as the state field, then pause program for half a second
                        if keyword_field == state_field:
                            print(f'{keyword_field} is the same as your state field.')
                            time.sleep(0.5)

                        # display that the keyword field is the same as the id field, then pause program for half a second
                        elif keyword_field == id_field:
                            print(f'{keyword_field} is the same as your id field.')
                            time.sleep(0.5)

                        # display that the field does not exist, then pause program for half a second
                        else:
                            print(f'{keyword_field} field does not exist.')
                            time.sleep(0.5)

                        # ask user again for keyword field
                        keyword_field = input('Enter the field that contains keywords: ')

                    # initialize list to store keyword values object containing values from keyword_field
                    keyword_data = []

                    # add data values/object of keyword_field to keyword_data list
                    keyword_data.append(contents[keyword_field])

                    # iterate through each keyword result object and add unique values to keyword list
                    for ob in keyword_data:
                        for word in ob:
                            # add keyword to keywords list if it isn't already in it
                            if word not in keywords:
                                keywords.append(word)

                    # handle empty files
                    if len(keywords) == 0:
                        print('No keywords were found.')

                        # return empty states, ids, and user_keywords lists to parent function
                        return [], [], []

                    # run this code if data is in the file
                    else:
                        time.sleep(1.5)     # pause program for a second and a half

                        # display menu of fields that were found in file
                        print()
                        print('Available Keywords')
                        print('------------------')
                        for i, keyword in enumerate(keywords, start=1):
                            print(f'{i}. {keyword}')
                        print()

                        time.sleep(1)   # pause program for a second

                        # determine how many keywords are available
                        total_keywords = len(keywords)

                        # initialize user_keywords list to store keywords user wants to pull data for
                        user_keywords = []

                        # initialize valid_number bool to False to use in below validation loop
                        valid_number = False

                        # validate the input is an integer and smaller than the total number of available keywords
                        while valid_number == False:
                            # ask user how many keywords they want to pull data for
                            num_keywords = input('How many keywords would you like to pull data for? ')

                            # tell user if input is not an integer
                            if not num_keywords.isdigit():
                                print(f'{num_keywords} is not an integer.')

                                time.sleep(0.5)     # pause program for half a second

                            # tell user if input is larger than total number of available keywords
                            elif int(num_keywords) > total_keywords:
                                print(f'{num_keywords} is larger than the amount of available keywords: {total_keywords}.')

                                time.sleep(0.5)     # pause program for half a second

                            # tell user if input equals the total number of available keywords
                            elif int(num_keywords) == total_keywords:
                                print(f'{num_keywords} is the same as the total number of available keywords: {total_keywords}.')

                                time.sleep(0.5)     # pause program for half a second

                            # tell user if input is equal to 0
                            elif int(num_keywords) == 0:
                                print(f'You need to use at least 1 keyword.')

                                time.sleep(0.5)     # pause program for half a second

                            # the input passed both validation tests
                            else:
                                # convert valid input to integer
                                num_keywords = int(num_keywords)

                                # signal that a valid input has been entered
                                valid_number = True

                        # new line
                        print()

                        # initialize counter
                        counter = 0

                        # get data for as many keywords as user wants
                        while counter < num_keywords:
                            # ask user to enter an available keyword from the menu
                            user_keyword = input(f'Enter keyword #{counter+1} from the menu: ')

                            # validate that the keyword is an option
                            while user_keyword not in keywords or user_keyword in user_keywords:
                                # tell user if keyword is not in keywords list
                                if user_keyword not in keywords:
                                    print(f'{user_keyword} is not an option. Please try again')

                                    time.sleep(0.5)     # pause program for half a second

                                # tell user if data for keyword has already been pulled
                                else:
                                    print(f'Data for {user_keyword} has already been pulled.')

                                    time.sleep(0.5)     # pause program for half a second

                                # new line
                                print()

                                # ask user again to enter an available keyword from the menu
                                user_keyword = input(f'Enter keyword #{counter+1} from the menu: ')

                            # new line
                            print()

                            # add keyword to user_keywords list
                            user_keywords.append(user_keyword)

                            # filter contents using user specified keyword
                            keyword_contents = contents[contents[keyword_field] == user_keyword]
                            
                            # get number of results pulled
                            num_results = len(keyword_contents)

                            # display results
                            if num_keywords > 1:
                                print(f'\nPulling data for {user_keyword} returned {num_results} results.\n')

                            time.sleep(1)   # pause program for a second
                            print(f'Results For - {user_keyword}:\n{keyword_contents}\n')

                            # add data values/object of state_field and id_field (after keyword filtering) to respective list
                            state_data.append(keyword_contents[state_field])
                            
                            id_data.append(keyword_contents[id_field])

                            # add 1 to the counter
                            counter += 1

                # run this code if user indicated that they do not want to filter file data using a keyword
                else:
                    # add data values/object of state_field and id_field (without keyword filtering) to respective list
                    state_data.append(contents[state_field])
                    contents[id_field] = contents[id_field].astype(float)
                    
                    id_data.append(contents[id_field])

                    # set user keyword to an empty list
                    user_keywords = []

            # run this code if function parameters indicate to not filter by keyword
            else:
                # add data values/object of state_field and id_field (without keyword filtering) to respective list
                state_data.append(contents[state_field])
                id_data.append(contents[id_field])

                # set user keyword to an empty list
                user_keywords = []

            # iterate through each state result object and add the state value to the states list
            for ob in state_data:
                for state in ob:
                    states.append(state)

            # iterate through each id result object and add the id value to the ids list
            for ob in id_data:
                for tweet in ob:
                    # had to format like this to prevent truncated values from being added to ids list
                    ids.append(int('{:.0f}'.format(tweet)))

            # return states, ids, and user_keyword to parent function
            return states, ids, user_keywords